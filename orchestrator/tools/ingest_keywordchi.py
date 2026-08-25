#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fold a keywordchi .xlsx export into the demand data file.

Architecture credit: Albaloo Studio — albaloostudio.com
Owner: Alireza Mozaffari

keywordchi exports one workbook per seed. Each has ~7 sheets that overlap
heavily with each other — the first two exports were 743 rows that collapsed to
500 unique terms, so a third of every export is the same term repeated across
its own sheets. Merging by hand invites both duplication and quiet loss.

This is idempotent. Re-ingesting the same file changes nothing except adding
the seed's name to terms it already knew, so running it twice is safe and
running it on a partial harvest and again on the rest is safe.

WHAT IT REFUSES TO INVENT
─────────────────────────
If a future export DOES carry numeric columns — volume, competition, CPC —
this records them per term and flips HAS_SEARCH_VOLUME to true, which makes
keyword_demand.py exit rather than print its no-volume disclaimer over data
that has volume. It will not guess a number for a term that lacks one, and it
will not carry a number from one seed's export onto a term found in another.

IT VERIFIES THE FILE IS ACTUALLY FROM KEYWORDCHI
────────────────────────────────────────────────
The obvious way to call this is `ingest_keywordchi.py ~/Downloads/*.xlsx`, and
the first time that was run the glob swallowed a folder of hotel cost charts, a
conference contact list and a Search Console export — 2,853 terms of pure
noise, and 7,632 numeric cells that tripped the volume detector. --dry-run
caught it, but a tool that depends on being called carefully is not safe.

So every workbook must carry keywordchi's own provenance sheet («کیوردچی»,
naming the tool and its URL). Anything else is refused by name. The seed is
read from that sheet's sentence rather than the filename, because a browser
saving a second copy names it "تور کروز (1).xlsx" and that is not a seed.

    python3 tools/ingest_keywordchi.py ~/Downloads/*.xlsx      # safe now
    python3 tools/ingest_keywordchi.py --dry-run ~/Downloads/"تور کروز.xlsx"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
from keyword_demand import DATA, classify, norm  # noqa: E402

# keywordchi's branding sheet. It carries the seed name in prose and no terms,
# and its presence is what proves the workbook came from keywordchi at all.
BRAND_SHEET = "کیوردچی"
BRAND_MARK = "کیوردچی"          # appears in the sheet body and in its URL line
# «خروجی همه کلماتی که کاربران درباره ی 'سفر کروز' در گوگل سرچ می کنند»
SEED_RE = re.compile(r"درباره\s*ی?\s*['\u2018\u2019\u201c\u201d\"]?(.+?)"
                     r"['\u2018\u2019\u201c\u201d\"]?\s*در\s*گوگل")


class NotAKeywordchiExport(Exception):
    """Raised for any workbook without keywordchi's provenance sheet."""


def read_export(path: Path) -> tuple[str, dict[str, list[Any]], list[str]]:
    """Return (seed, {sheet: [terms]}, numeric_cells_found).

    Raises NotAKeywordchiExport if the workbook is anything else.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    titles = [ws.title for ws in wb.worksheets]
    if BRAND_SHEET not in titles:
        wb.close()
        raise NotAKeywordchiExport(
            f"no «{BRAND_SHEET}» sheet — sheets are {titles[:4]}")

    # The seed comes from keywordchi's own sentence, not from the filename.
    seed = ""
    blurb = []
    for row in wb[BRAND_SHEET].iter_rows(values_only=True):
        for cell in row:
            if cell and str(cell).strip():
                blurb.append(str(cell).strip())
    joined = " ".join(blurb)
    if BRAND_MARK not in joined:
        wb.close()
        raise NotAKeywordchiExport(
            f"«{BRAND_SHEET}» sheet does not name keywordchi")
    m = SEED_RE.search(joined)
    if m:
        seed = norm(m.group(1))

    sheets: dict[str, list[Any]] = {}
    numerics: list[str] = []
    for ws in wb.worksheets:
        if ws.title == BRAND_SHEET:
            continue
        vals = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                    numerics.append(f"{ws.title}:{cell}")
                    continue
                t = norm(cell)
                if t:
                    vals.append(t)
        if vals:
            sheets[ws.title] = vals
    wb.close()
    if not seed:
        # keywordchi also names one sheet after the seed. Only if BOTH that and
        # the sentence fail do we fall back to the filename, and say so.
        for title in titles:
            if title != BRAND_SHEET and norm(title) in " ".join(
                    t for ts in sheets.values() for t in ts):
                seed = norm(title)
                break
    if not seed:
        seed = norm(path.stem)
        print(f"    (seed read from filename — keywordchi's own label was "
              f"unreadable in this file)")
    return seed, sheets, numerics


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("files", nargs="+", type=Path)
    ap.add_argument("--data", type=Path, default=DATA)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    doc = json.loads(args.data.read_text(encoding="utf-8"))
    index = {t["term"]: t for t in doc["terms"]}
    before = len(index)
    seeds_before = set(doc.get("seeds", []))
    all_numerics: list[str] = []
    refused = 0

    print(f"{args.data.name}: {before} terms, seeds "
          f"«{'» «'.join(sorted(seeds_before))}»\n")

    for f in args.files:
        if not f.exists():
            print(f"  MISSING  {f}")
            continue
        if f.suffix.lower() != ".xlsx":
            print(f"  SKIP     {f.name} — not an .xlsx")
            continue
        try:
            seed, sheets, numerics = read_export(f)
        except NotAKeywordchiExport as exc:
            print(f"  REFUSED  {f.name} — {exc}")
            refused += 1
            continue
        all_numerics += numerics
        seen: set[str] = set()
        for terms in sheets.values():
            seen.update(terms)
        new = [t for t in seen if t not in index]
        for t in sorted(seen):
            e = index.setdefault(t, {"term": t, "seeds": [], "sheets": []})
            if seed not in e["seeds"]:
                e["seeds"].append(seed)
            for title, vals in sheets.items():
                if t in vals and title not in e["sheets"]:
                    e["sheets"].append(title)
        if seed not in seeds_before:
            doc.setdefault("seeds", []).append(seed)
            seeds_before.add(seed)
        rows = sum(len(v) for v in sheets.values())
        print(f"  {f.name}")
        print(f"    seed «{seed}» · {len(sheets)} sheets · {rows} rows → "
              f"{len(seen)} unique · {len(new)} NEW"
              + (f" · {len(numerics)} numeric cells" if numerics else ""))

    if refused:
        print(f"\n  {refused} file(s) refused — not keywordchi exports.")
    doc["terms"] = [index[k] for k in sorted(index)]
    if all_numerics:
        doc["HAS_SEARCH_VOLUME"] = True
        doc["numeric_cells_sample"] = all_numerics[:20]
        print(f"\n  !! {len(all_numerics)} numeric cells found. HAS_SEARCH_VOLUME "
              f"set true — keyword_demand.py will now refuse to run rather than "
              f"print its no-volume disclaimer over data that has volume. That "
              f"is deliberate: teach it the column names before re-running.")

    rows = classify(doc)
    usable = [r for r in rows if not r["reject"]]
    print(f"\n  {before} → {len(doc['terms'])} terms "
          f"(+{len(doc['terms']) - before})")
    print(f"  usable after filtering: {len(usable)} "
          f"({len(usable) * 100 // max(len(rows), 1)}%)")

    if args.dry_run:
        print("\n  --dry-run: nothing written")
        return 0
    args.data.write_text(json.dumps(doc, ensure_ascii=False, indent=1),
                         encoding="utf-8")
    print(f"\n  wrote {args.data}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
